import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import PatternFill
import os
import re
import math
import traceback

# Définir les constantes
SHEET_NAME = "Annexe 2.1+MO-autorisation24"
PATIENT_COUNT_CELL = "B10"
COL_DESIGNATION = 1  # A
COL_LIMITE_OCCURRENCE = 2  # B
COL_COUT_SURCOUT = 3  # C
COL_MONTANT_UNITAIRE = 4  # D
COL_NOMBRE_ITEMS = 5  # E
COL_TOTAL_LIGNE = 6  # F
COL_TOTAL_CENTRE = 7  # G
COL_CONSIGNES = 8  # H
START_ROW = 18
END_ROW_MARKER = "TOTAL GÉNÉRAL"
TOTAL_GENERAL_ROW_OFFSET = 1
TOLERANCE = 0.01

# Couleurs de surlignage
HIGHLIGHT_COLOR_DEFAULT = "ADD8E6"  # Bleu clair (par défaut)
HIGHLIGHT_COLOR_LEVEL = "FFC7CE"    # Rouge clair (pour lignes spécifiques au niveau)
NO_FILL = PatternFill(fill_type=None) # Pour effacer le remplissage

# Coûts horaires par niveau et par type de ligne
COUT_HORAIRE = {
    "screening": { "1": 57.5, "2": 115.0, "3": 172.5 },
    "visite_site": { "1": 57.5, "2": 115.0, "3": 115.0 },
    "visite_finale": { "1": 57.5, "2": 115.0, "3": 115.0 }
}

# Temps de base par niveau et par type de ligne
TEMPS_BASE = {
    "screening": { "1": 1.0, "2": 2.0, "3": 3.0 },
    "visite_site": { "1": 1.0, "2": 2.0, "3": 2.0 },
    "visite_finale": { "1": 1.0, "2": 2.0, "3": 2.0 }
}

def extract_max_hour(text):
    if not isinstance(text, str): return 0.0
    numbers = re.findall(r"\d+[.,]?\d*", text.replace(",", "."))
    max_h = 0.0
    for num_str in numbers:
        try:
            h = float(num_str)
            if h > max_h: max_h = h
        except ValueError: continue
    return max_h

def safe_float(value, default=0.0):
    if value is None: return default
    try:
        if isinstance(value, str) and ('%' in value or value.strip() == ''): return default
        if isinstance(value, str): value = value.replace(",", ".").strip()
        return float(value)
    except (ValueError, TypeError): return default

def extract_montants_par_niveau(text):
    if not isinstance(text, str): return {}
    montants = {}
    patterns = [r"niveau\s*1\s*:?\s*(\d+[.,]?\d*)", r"niveau\s*2\s*:?\s*(\d+[.,]?\d*)", r"niveau\s*3\s*:?\s*(\d+[.,]?\d*)"]
    for i, pattern in enumerate(patterns, 1):
        matches = re.findall(pattern, text.lower())
        if matches:
            montants_niveau = [safe_float(m.replace(",", ".")) for m in matches]
            montants[str(i)] = max(montants_niveau)
    return montants

def extract_montants_par_centre(text):
    if not isinstance(text, str): return {}
    montants = {}
    patterns = [r"coordonnateur\s*:?\s*(\d+[.,]?\d*)", r"associé\s*:?\s*(\d+[.,]?\d*)"]
    centres = ["Coordonnateur", "Associé"]
    for i, pattern in enumerate(patterns):
        matches = re.findall(pattern, text.lower())
        if matches:
            montants_centre = [safe_float(m.replace(",", ".")) for m in matches]
            montants[centres[i]] = max(montants_centre)
    return montants

def extract_time_hours(text):
    if not isinstance(text, str): return 0.0
    hour_patterns = [r"(\d+[.,]?\d*)\s*h(?:eures?)?", r"(\d+[.,]?\d*)\s*heure(?:s)?"]
    for pattern in hour_patterns:
        matches = re.findall(pattern, text.lower())
        if matches: return safe_float(matches[0].replace(",", "."))
    minute_patterns = [r"(\d+[.,]?\d*)\s*min(?:utes?)?", r"(\d+[.,]?\d*)\s*minute(?:s)?"]
    for pattern in minute_patterns:
        matches = re.findall(pattern, text.lower())
        if matches: return safe_float(matches[0].replace(",", ".")) / 60.0
    return 0.0

def calculate_additional_time(study_level, num_pages_crf):
    if study_level == "1": return (num_pages_crf // 10) * 0.25
    elif study_level in ["2", "3"]: return (num_pages_crf // 5) * 0.25
    return 0.0


class MatriceApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Assistant Matrice Coûts v4.3 (Scrollable)")
        self.master.geometry("600x800")
        self.master.minsize(550, 600)

        self.ICON_GENERATE_B64 = "R0lGODlhEAAQALMAAF9bWx8fH7+/vzc3N4mJiWlpaa+vrzAwMMbGxmdnZz4+Pujo6P///wAAAAAAAP///yH5BAEAAA8ALAAAAAAQABAAAARa8EnJpAKgKz5s759AAMGAgoMWcZJpCCaFAhQBHIiAYEigpEAkDE6lBIrG4/F4alEqrVI9mUqvV8uEKkCAEBgMBAICpRAJA4EgUGoNCpfd7fa8fgEAOw=="
        self.ICON_CLEAR_B64 = "R0lGODlhEAAQALMAAF9bWx8fH7+/vzc3N4mJiWlpaa+vrzAwMMbGxmdnZz4+Pujo6P///wAAAAAAAP///yH5BAEAAA8ALAAAAAAQABAAAARe8MkJgC2goktY758BEEBAgYJgjCRpFBqFAZUAaLgCCARFo7FoPC6YDw6n0qmYdKr1esdqtVqv14sCCBAgAAiCQyAQGAqEQyAQGAqG4/F4fP7/gYIAOw=="
        self.ICON_QUIT_B64 = "R0lGODlhEAAQALMAAF9bWx8fH7+/vzc3N4mJiWlpaa+vrzAwMMbGxmdnZz4+Pujo6P///wAAAAAAAP///yH5BAEAAA8ALAAAAAAQABAAAARa8MkJgC0gIjls758DEEAgYGCcJGkUGoUBlQBQuAIIRMWjsWg8LpgPDqfSqZh0qtV6x2q1Wq/XiwIIECAACIJDIBAYCoRDIBCICoaj8Xh8/v+BggA7"

        self._setup_styles()
        self._create_scrollable_container()
        self._create_widgets()
        self._layout_widgets()

    def _setup_styles(self):
        self.style = ttk.Style(self.master)
        self.style.theme_use('clam')
        self.generate_icon = tk.PhotoImage(data=self.ICON_GENERATE_B64)
        self.clear_icon = tk.PhotoImage(data=self.ICON_CLEAR_B64)
        self.quit_icon = tk.PhotoImage(data=self.ICON_QUIT_B64)
        try:
            default_font = ('Segoe UI', 10)
            self.master.option_add("*Font", default_font)
        except tk.TclError:
            pass
        self.style.configure("TLabel", padding=5)
        self.style.configure("TEntry", padding=5)
        self.style.configure("TCombobox", padding=5)
        self.style.configure("TCheckbutton", padding=(10, 5))
        self.style.configure("TButton", padding=8, font=('Segoe UI', 10, 'bold'))
        self.style.configure("TLabelframe.Label", font=('Segoe UI', 11, 'bold'), padding=(0,0,0,5))
        self.style.configure("Accent.TButton", background="#0078D7", foreground="white")
        self.style.map("Accent.TButton", background=[('active', '#005a9e'), ('pressed', '!disabled', '#004c8c')])

    def _create_scrollable_container(self):
        container = ttk.Frame(self.master)
        container.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(container, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas, padding="20 20 20 20")

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
    def _create_widgets(self):
        # Tous les widgets sont maintenant des enfants de 'self.scrollable_frame'
        self.title_label = ttk.Label(self.scrollable_frame, text="Assistant de Remplissage de Matrice", font=('Segoe UI', 14, 'bold'), anchor="center")
        self.intro_label = ttk.Label(self.scrollable_frame, text="Remplissez les champs ci-dessous et cliquez sur 'Générer' pour créer votre fichier.", wraplength=480, anchor="center")
        
        # --- Cadre des paramètres ---
        self.params_frame = ttk.LabelFrame(self.scrollable_frame, text="Paramètres de l'étude", padding="15")
        self.niveau_var = tk.StringVar()
        self.niveau_combo = ttk.Combobox(self.params_frame, textvariable=self.niveau_var, values=["1", "2", "3"], state="readonly")
        self.patients_var = tk.StringVar()
        self.patients_entry = ttk.Entry(self.params_frame, textvariable=self.patients_var)
        self.visites_var = tk.StringVar()
        self.visites_entry = ttk.Entry(self.params_frame, textvariable=self.visites_var)
        self.centre_var = tk.StringVar()
        self.centre_combo = ttk.Combobox(self.params_frame, textvariable=self.centre_var, values=["Coordonnateur", "Associé"], state="readonly")
        self.duree_var = tk.StringVar()
        self.duree_entry = ttk.Entry(self.params_frame, textvariable=self.duree_var)
        self.pages_crf_var = tk.StringVar(value="0")
        self.pages_crf_entry = ttk.Entry(self.params_frame, textvariable=self.pages_crf_var)

        # --- Cadre d'estimation infirmier ---
        self.infirmier_frame = ttk.LabelFrame(self.scrollable_frame, text="Estimation du temps infirmier (optionnel)", padding="15")
        self.prelevements_sang_var = tk.StringVar(value="")
        self.prelevements_sang_entry = ttk.Entry(self.infirmier_frame, textvariable=self.prelevements_sang_var)
        self.prelevements_urine_var = tk.StringVar(value="")
        self.prelevements_urine_entry = ttk.Entry(self.infirmier_frame, textvariable=self.prelevements_urine_var)
        self.signes_vitaux_var = tk.StringVar(value="")
        self.signes_vitaux_entry = ttk.Entry(self.infirmier_frame, textvariable=self.signes_vitaux_var)
        self.injections_var = tk.StringVar(value="")
        self.injections_entry = ttk.Entry(self.infirmier_frame, textvariable=self.injections_var)
        self.perfusions_var = tk.StringVar(value="")
        self.perfusions_entry = ttk.Entry(self.infirmier_frame, textvariable=self.perfusions_var)
        self.catheters_var = tk.StringVar(value="")
        self.catheters_entry = ttk.Entry(self.infirmier_frame, textvariable=self.catheters_var)
        self.pk_pd_var = tk.StringVar(value="")
        self.pk_pd_entry = ttk.Entry(self.infirmier_frame, textvariable=self.pk_pd_var)
        
        # --- Cadre des options ---
        self.options_frame = ttk.LabelFrame(self.scrollable_frame, text="Options", padding="15")
        self.avenants_var = tk.StringVar(value="0")
        self.avenants_entry = ttk.Entry(self.options_frame, textvariable=self.avenants_var)
        self.monitoring_var = tk.StringVar(value="0")
        self.monitoring_entry = ttk.Entry(self.options_frame, textvariable=self.monitoring_var)
        self.auto_q_count_var = tk.StringVar(value="")
        self.auto_q_count_entry = ttk.Entry(self.options_frame, textvariable=self.auto_q_count_var)
        self.auto_q_format_var = tk.StringVar(value="électronique")
        self.auto_q_format_combo = ttk.Combobox(self.options_frame, textvariable=self.auto_q_format_var, values=["papier", "électronique"], state="readonly")
        self.personnel_var = tk.BooleanVar()
        self.personnel_check = ttk.Checkbutton(self.options_frame, text="Personnel extérieur", variable=self.personnel_var)

        # --- Cadre des boutons ---
        self.button_frame = ttk.Frame(self.scrollable_frame)
        self.generate_button = ttk.Button(self.button_frame, text="Générer/MàJ Matrice", command=self._generate_matrix_wrapper, style="Accent.TButton", image=self.generate_icon, compound=tk.LEFT)
        self.clear_button = ttk.Button(self.button_frame, text="Effacer", command=self.clear_quantities, image=self.clear_icon, compound=tk.LEFT)
        self.quit_button = ttk.Button(self.button_frame, text="Quitter", command=self.master.quit, image=self.quit_icon, compound=tk.LEFT)

    def _layout_widgets(self):
        # Configure le 'scrollable_frame' pour qu'il s'étende horizontalement
        self.scrollable_frame.columnconfigure(0, weight=1)

        self.title_label.grid(row=0, column=0, pady=(0, 5), sticky="ew")
        self.intro_label.grid(row=1, column=0, pady=(0, 15), sticky="ew")

        # Layout paramètres
        self.params_frame.grid(row=2, column=0, pady=5, sticky="ew")
        self.params_frame.columnconfigure(1, weight=1)
        labels_params = ["Niveau de l'étude:", "Nombre de patients:", "Nb total visites/patient:", "Type de centre:", "Durée étude (années):", "Nombre de pages CRF:"]
        widgets_params = [self.niveau_combo, self.patients_entry, self.visites_entry, self.centre_combo, self.duree_entry, self.pages_crf_entry]
        for i, (label_text, widget) in enumerate(zip(labels_params, widgets_params)):
            ttk.Label(self.params_frame, text=label_text).grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            widget.grid(row=i, column=1, sticky=tk.EW, padx=5, pady=2)

        # Layout estimation infirmier
        self.infirmier_frame.grid(row=3, column=0, pady=10, sticky="ew")
        self.infirmier_frame.columnconfigure(1, weight=1)
        labels_infirmier = ["Nb prélèvements sanguins:", "Nb prélèvements d'urine:", "Nb mesures signes vitaux:", "Nb injections traitement:", "Nb pose/retrait perfusion:", "Nb pose/retrait cathéter:", "Nb points PK/PD:"]
        widgets_infirmier = [self.prelevements_sang_entry, self.prelevements_urine_entry, self.signes_vitaux_entry, self.injections_entry, self.perfusions_entry, self.catheters_entry, self.pk_pd_entry]
        for i, (label_text, widget) in enumerate(zip(labels_infirmier, widgets_infirmier)):
            ttk.Label(self.infirmier_frame, text=label_text).grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            widget.grid(row=i, column=1, sticky=tk.EW, padx=5, pady=2)

        # Layout options
        self.options_frame.grid(row=4, column=0, pady=5, sticky="ew")
        self.options_frame.columnconfigure(1, weight=1)
        labels_options = ["Nb avenants:", "Nb visites monitoring:", "Nb auto-questionnaires:", "Format auto-questionnaire:"]
        widgets_options = [self.avenants_entry, self.monitoring_entry, self.auto_q_count_entry, self.auto_q_format_combo]
        for i, (label_text, widget) in enumerate(zip(labels_options, widgets_options)):
            ttk.Label(self.options_frame, text=label_text).grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            widget.grid(row=i, column=1, sticky=tk.EW, padx=5, pady=2)
        self.personnel_check.grid(row=len(labels_options), column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)

        # Layout boutons
        self.button_frame.grid(row=5, column=0, pady=(20, 0), sticky="ew")
        self.button_frame.columnconfigure((0, 2), weight=1)
        self.generate_button.grid(row=0, column=0, padx=5, sticky="e")
        self.clear_button.grid(row=0, column=1, padx=5)
        self.quit_button.grid(row=0, column=2, padx=5, sticky="w")
    
    def validate_inputs(self):
        errors = []
        try:
            if not self.niveau_var.get(): errors.append("Niveau de l'étude manquant.")
            if not self.patients_var.get() or int(self.patients_var.get()) <= 0: errors.append("Nombre de patients invalide (> 0).")
            if not self.visites_var.get() or int(self.visites_var.get()) <= 0: errors.append("Nombre de visites invalide (> 0).")
            if not self.centre_var.get(): errors.append("Type de centre manquant.")
            if not self.duree_var.get() or int(self.duree_var.get()) <= 0: errors.append("Durée d'étude invalide (> 0).")
            if not self.avenants_var.get() or int(self.avenants_var.get()) < 0: errors.append("Nombre d'avenants invalide (>= 0).")
            if not self.monitoring_var.get() or int(self.monitoring_var.get()) < 0: errors.append("Nombre de visites de monitoring invalide (>= 0).")
            if not self.pages_crf_var.get() or int(self.pages_crf_var.get()) < 0: errors.append("Nombre de pages CRF invalide (>= 0).")
            if self.auto_q_count_var.get() and int(self.auto_q_count_var.get()) < 0:
                errors.append("Nombre d'auto-questionnaires invalide (>= 0).")
            if int(self.visites_var.get()) < 2:
                errors.append("Le nombre total de visites doit être au moins 2 (1 screening + 1 finale).")
        except ValueError:
            errors.append("Veuillez saisir des nombres valides pour tous les champs requis.")
        except Exception as e:
             errors.append(f"Erreur inattendue dans les saisies : {e}")
        if errors:
            messagebox.showerror("Erreur de saisie", "Veuillez corriger les erreurs suivantes:\n- " + "\n- ".join(errors))
            return False
        return True

    def _generate_matrix_wrapper(self):
        if not self.validate_inputs(): return
        source_file = filedialog.askopenfilename(title="Sélectionner le fichier matrice Excel modèle (.xlsm)", filetypes=[("Fichiers Excel", "*.xlsm")])
        if not source_file: return
        output_file = filedialog.asksaveasfilename(title="Enregistrer la matrice remplie sous (.xlsm)", defaultextension=".xlsm", filetypes=[("Fichiers Excel", "*.xlsm")], initialfile="matrice_remplie.xlsm")
        if not output_file: return
        try:
            self.generate_matrix_logic(source_file, output_file)
        except KeyError:
             messagebox.showerror("Erreur", f"La feuille '{SHEET_NAME}' est introuvable dans le fichier sélectionné.")
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de la génération de la matrice :\n{type(e).__name__}: {e}")

    def generate_matrix_logic(self, source_file, output_file):
        workbook = openpyxl.load_workbook(source_file, keep_vba=True)
        sheet = workbook[SHEET_NAME]

        studyLevel = self.niveau_var.get()
        numPatients = int(self.patients_var.get())
        numVisits = int(self.visites_var.get())
        centerType = self.centre_var.get()
        studyDuration = int(self.duree_var.get())
        numAvenants = int(self.avenants_var.get())
        numVisitesMonitoring = int(self.monitoring_var.get())
        personnelExterieur = self.personnel_var.get()
        numPagesCRF = int(self.pages_crf_var.get())
        numAutoQ = int(self.auto_q_count_var.get()) if self.auto_q_count_var.get() else 0
        formatAutoQ = self.auto_q_format_var.get()

        def get_optional_int(var):
            val = var.get()
            return int(val) if val.isdigit() else None
        
        nb_prelevements_sang = get_optional_int(self.prelevements_sang_var)
        nb_prelevements_urine = get_optional_int(self.prelevements_urine_var)
        nb_signes_vitaux = get_optional_int(self.signes_vitaux_var)
        nb_injections = get_optional_int(self.injections_var)
        nb_perfusions = get_optional_int(self.perfusions_var)
        nb_catheters = get_optional_int(self.catheters_var)
        nb_pk_pd = get_optional_int(self.pk_pd_var)

        numVisitsSurSite = max(0, numVisits - 2)

        highlight_fill_default = PatternFill(start_color=HIGHLIGHT_COLOR_DEFAULT, end_color=HIGHLIGHT_COLOR_DEFAULT, fill_type="solid")
        highlight_fill_level = PatternFill(start_color=HIGHLIGHT_COLOR_LEVEL, end_color=HIGHLIGHT_COLOR_LEVEL, fill_type="solid")

        sheet[PATIENT_COUNT_CELL].value = numPatients
        sheet[PATIENT_COUNT_CELL].fill = highlight_fill_default
        
        firstRow, lastRow, totalRow = self.find_data_rows(sheet)
        if not (firstRow > 0 and lastRow >= firstRow):
             raise ValueError("Impossible de déterminer la plage de données de la matrice.")

        total_general = 0.0
        processed_rows = []
        
        for r in range(firstRow, lastRow + 1):
            designation_cell = sheet.cell(row=r, column=COL_DESIGNATION)
            designation = str(designation_cell.value).strip() if designation_cell.value is not None else ""
            if not designation: continue

            quantity_per_patient_or_center, montant_unitaire, is_level_specific, is_center_specific, is_fixed_cost_line, is_special_calculation, special_calc_key = None, None, False, False, False, False, ""
            designation_lower = designation.lower()
            
            # --- LOGIQUE DE CALCUL PAR LIGNE ---

            if "frais administratifs" in designation_lower:
                quantity_per_patient_or_center, is_center_specific, is_fixed_cost_line = 1, True, True
            elif "frais supplémentaires pour l'élaboration d'un avenant" in designation_lower:
                if numAvenants > 0: quantity_per_patient_or_center, is_center_specific, is_fixed_cost_line = numAvenants, True, True
            elif "mise en place de la recherche" in designation_lower:
                quantity_per_patient_or_center, is_level_specific, is_fixed_cost_line = 1, True, True
            elif "forfait de frais logistique" in designation_lower:
                is_personnel_exterieur_line = "personnels extérieurs" in designation_lower
                if not is_personnel_exterieur_line or (is_personnel_exterieur_line and personnelExterieur): quantity_per_patient_or_center, is_level_specific = numVisits, True
            elif "forfait maintenance des appareils" in designation_lower:
                quantity_per_patient_or_center, is_fixed_cost_line = studyDuration, True
            elif "consultation d'inclusion" in designation_lower:
                quantity_per_patient_or_center, is_level_specific = 1, True
            elif "prise de connaissance de l'amendement" in designation_lower or "prise de connaissance de l'addendum" in designation_lower:
                if numAvenants > 0:
                    time_per_amendment = extract_time_hours(designation) or extract_time_hours(str(sheet.cell(row=r, column=COL_CONSIGNES).value))
                    quantity_per_patient_or_center, is_fixed_cost_line = numAvenants * (time_per_amendment or 0.5), True
            elif "consultation pour addendum" in designation_lower or "consultation pour amendement" in designation_lower:
                if numAvenants > 0:
                    time_per_addendum = extract_time_hours(designation) or extract_time_hours(str(sheet.cell(row=r, column=COL_CONSIGNES).value))
                    quantity_per_patient_or_center, is_fixed_cost_line = numAvenants * (time_per_addendum or 1.0), True
            elif "temps tec formation" in designation_lower and not "questionnaires" in designation_lower:
                if "niveau 1" in designation_lower and studyLevel == "1": quantity_per_patient_or_center = 5
                if "niveau 2" in designation_lower and studyLevel == "2": quantity_per_patient_or_center = 6
                if "niveau 3" in designation_lower and studyLevel == "3": quantity_per_patient_or_center = 8
                if quantity_per_patient_or_center is not None: is_fixed_cost_line, is_level_specific = True, True
            elif "temps tec monitoring avec promoteur/cro" in designation_lower:
                hrs_per_visit = 0
                if "niveau 1" in designation_lower and studyLevel == "1": hrs_per_visit = 2.5
                if "niveau 2" in designation_lower and studyLevel == "2": hrs_per_visit = 4
                if "niveau 3" in designation_lower and studyLevel == "3": hrs_per_visit = 5
                if hrs_per_visit > 0: quantity_per_patient_or_center, is_fixed_cost_line, is_level_specific = numVisitesMonitoring * hrs_per_visit, True, True
            elif "temps tec visite de screening patient" in designation_lower:
                is_special_calculation, is_level_specific, quantity_per_patient_or_center, special_calc_key = True, True, 1, "screening"
            elif "temps tec visite sur site, de suivi patient ou téléphonique" in designation_lower:
                is_special_calculation, is_level_specific, quantity_per_patient_or_center, special_calc_key = True, True, numVisitsSurSite, "visite_site"
            elif "temps tec visite finale ou arrêt prématuré" in designation_lower:
                is_special_calculation, is_level_specific, quantity_per_patient_or_center, special_calc_key = True, True, 1, "visite_finale"
            elif "temps tec formation aux questionnaires et carnets patient" in designation_lower:
                quantity_per_patient_or_center, montant_unitaire, is_fixed_cost_line = 1, 57.50, True
            elif "temps tec gestion auto-questionnaire" in designation_lower:
                quantity_per_patient_or_center, is_fixed_cost_line = numAutoQ, False
                montant_unitaire = 28.75 if numAutoQ > 5 else 14.37
            elif "temps tec formation initiale du patient à l'auto-questionnaire" in designation_lower:
                quantity_per_patient_or_center, is_fixed_cost_line = 1, False
                montant_unitaire = (86.25 if formatAutoQ == "électronique" else 43.12) if numAutoQ > 5 else (57.5 if formatAutoQ == "électronique" else 28.75)
            elif "temps tec pour la gestion des kits de prélèvement" in designation_lower:
                quantity_per_patient_or_center, montant_unitaire, is_fixed_cost_line = numVisits, 57.50, False
            elif "temps tec appel ivrs/iwrs" in designation_lower:
                quantity_per_patient_or_center, montant_unitaire, is_fixed_cost_line = numVisits, 11.24, False
            elif "temps tec pour la gestion des remboursements des frais patients" in designation_lower:
                quantity_per_patient_or_center, is_fixed_cost_line = numVisits, False
                montant_unitaire = 47.92 if "47,92" in str(sheet.cell(row=r, column=COL_MONTANT_UNITAIRE).value) else 19.17
            elif re.search(r"temps\s+ide\s*:\s*formation\s+au\s+protocole\s+initial", designation_lower):
                quantity_per_patient_or_center, is_fixed_cost_line, is_level_specific = 1, True, True
            elif "temps infirmier pour prélèvements sanguins" in designation_lower:
                quantity_per_patient_or_center = nb_prelevements_sang if nb_prelevements_sang is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 13.00, False
            elif "temps infirmier pour prélèvements d'urine" in designation_lower:
                quantity_per_patient_or_center = nb_prelevements_urine if nb_prelevements_urine is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 13.00, False
            elif "temps infirmier pour la mesure des signes vitaux" in designation_lower:
                quantity_per_patient_or_center = nb_signes_vitaux if nb_signes_vitaux is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 13.00, False
            elif re.search(r"temps\s+infirmier.*injection.*traitement", designation_lower):
                quantity_per_patient_or_center = nb_injections if nb_injections is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 13.00, False
            elif re.search(r"temps\s+infirmier.*pose.*retrait.*perfusion", designation_lower):
                quantity_per_patient_or_center = nb_perfusions if nb_perfusions is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 26.00, False
            elif re.search(r"temps\s+infirmier.*pose.*retrait.*cathéter", designation_lower):
                quantity_per_patient_or_center = nb_catheters if nb_catheters is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 26.00, False
            elif re.search(r"temps\s+infirmier.*aide\s+au\s+médecin", designation_lower):
                quantity_per_patient_or_center, is_fixed_cost_line = numVisits, False
            elif re.search(r"temps\s+infirmier.*point\s+de\s+pk/pd", designation_lower):
                quantity_per_patient_or_center = nb_pk_pd if nb_pk_pd is not None else numVisits
                montant_unitaire, is_fixed_cost_line = 13.00, False
            elif re.search(r"temps\s+manipulateur\s+radio.*administration", designation_lower):
                quantity_per_patient_or_center, montant_unitaire, is_fixed_cost_line = numVisits, 28.75, False

            # --- CALCULS GÉNÉRIQUES ---
            if quantity_per_patient_or_center is not None and quantity_per_patient_or_center >= 0:
                montant_unitaire_cell = sheet.cell(row=r, column=COL_MONTANT_UNITAIRE)
                if montant_unitaire is None: montant_unitaire = safe_float(montant_unitaire_cell.value)
                if is_level_specific and not is_special_calculation:
                    montants_niveau = extract_montants_par_niveau(str(montant_unitaire_cell.value))
                    if studyLevel in montants_niveau: montant_unitaire = montants_niveau[studyLevel]
                elif is_center_specific:
                    montants_centre = extract_montants_par_centre(str(montant_unitaire_cell.value))
                    if centerType in montants_centre: montant_unitaire = montants_centre[centerType]
                if is_special_calculation:
                    base_time = TEMPS_BASE[special_calc_key][studyLevel]
                    additional_time = calculate_additional_time(studyLevel, numPagesCRF)
                    total_time_per_visit = base_time + additional_time
                    montant_unitaire = total_time_per_visit * COUT_HORAIRE[special_calc_key][studyLevel]
                total_ligne = quantity_per_patient_or_center * montant_unitaire
                total_centre = total_ligne if is_fixed_cost_line else (quantity_per_patient_or_center * montant_unitaire * numPatients)
                sheet.cell(row=r, column=COL_NOMBRE_ITEMS).value = quantity_per_patient_or_center
                sheet.cell(row=r, column=COL_TOTAL_LIGNE).value = total_ligne
                sheet.cell(row=r, column=COL_TOTAL_CENTRE).value = total_centre
                fill_color = highlight_fill_level if is_level_specific or is_center_specific or is_special_calculation else highlight_fill_default
                for c in range(COL_MONTANT_UNITAIRE, COL_TOTAL_CENTRE + 1):
                    sheet.cell(row=r, column=c).fill = fill_color
                total_general += total_centre
                ligne_type = special_calc_key if is_special_calculation else "autre"
                processed_rows.append({
                    "designation": designation,
                    "type": ligne_type,
                    "quantite": quantity_per_patient_or_center,
                    "total_ligne": total_ligne,
                    "total_centre": total_centre,
                })
        if totalRow > 0:
            sheet.cell(row=totalRow, column=COL_TOTAL_CENTRE).value = total_general
            sheet.cell(row=totalRow, column=COL_TOTAL_CENTRE).fill = highlight_fill_default
        workbook.save(output_file)
        lignes_modifiees = len(processed_rows)
        visites_counts = {"screening": 0, "visite_site": 0, "visite_finale": 0}
        for row_info in processed_rows:
            if row_info["type"] in visites_counts:
                visites_counts[row_info["type"]] += 1
        messagebox.showinfo(
            "Succès",
            (
                "Matrice générée avec succès.\n"
                f"Fichier enregistré : {os.path.basename(output_file)}\n"
                f"Lignes mises à jour : {lignes_modifiees}\n"
                "Répartition des visites : "
                f"screening {visites_counts['screening']}, "
                f"visite sur site {visites_counts['visite_site']}, "
                f"visite finale {visites_counts['visite_finale']}\n"
                f"Total général : {total_general:.2f}"
            ),
        )

    def find_data_rows(self, sheet):
        firstRow, lastRow, totalRow = 0, 0, 0
        foundStart = False
        for r in range(START_ROW, sheet.max_row + 1):
            designation_cell = sheet.cell(row=r, column=COL_DESIGNATION)
            designation = str(designation_cell.value) if designation_cell.value else ""
            if not foundStart and designation: firstRow, foundStart = r, True
            if foundStart and END_ROW_MARKER.lower() in designation.lower():
                lastRow = r - 1
                for i in range(r, r + 6):
                    total_text_cell = sheet.cell(row=i, column=COL_DESIGNATION)
                    total_text = str(total_text_cell.value).lower() if total_text_cell.value else ""
                    if "total" in total_text and "général" in total_text:
                        totalRow = i
                        break
                break
        if firstRow > 0 and not totalRow: lastRow = sheet.max_row
        return firstRow, lastRow, totalRow

    def clear_quantities(self):
        target_file = filedialog.askopenfilename(title="Sélectionner le fichier matrice Excel à effacer (.xlsm)", filetypes=[("Fichiers Excel", "*.xlsm")])
        if not target_file: return
        if not messagebox.askyesno("Confirmation", "Voulez-vous vraiment effacer toutes les quantités et calculs (colonnes E, F, G) et le total général de ce fichier ?\nLe fichier sera modifié directement."): return
        try:
            workbook = openpyxl.load_workbook(target_file, keep_vba=True)
            sheet = workbook[SHEET_NAME]
            firstRow, lastRow, totalRow = self.find_data_rows(sheet)
            if firstRow > 0 and lastRow >= firstRow:
                count_cleared = 0
                for r in range(firstRow, lastRow + 20): 
                    for c in [COL_NOMBRE_ITEMS, COL_TOTAL_LIGNE, COL_TOTAL_CENTRE]:
                        cell_to_clear = sheet.cell(row=r, column=c)
                        if cell_to_clear.value is not None or cell_to_clear.fill.fgColor.rgb != '00000000':
                            cell_to_clear.value = None
                            cell_to_clear.fill = NO_FILL
                            count_cleared += 1
                if totalRow > 0:
                    total_cell = sheet.cell(row=totalRow, column=COL_TOTAL_CENTRE)
                    if total_cell.value is not None or total_cell.fill.fgColor.rgb != '00000000':
                        total_cell.value = None
                        total_cell.fill = NO_FILL
                        count_cleared += 1
                b10_cell = sheet[PATIENT_COUNT_CELL]
                if b10_cell.value is not None or b10_cell.fill.fgColor.rgb != '00000000':
                    b10_cell.value = None
                    b10_cell.fill = NO_FILL
                    count_cleared += 1
                workbook.save(target_file)
                messagebox.showinfo("Succès", f"{count_cleared} cellule(s) ont été effacées dans {os.path.basename(target_file)}.")
            else: messagebox.showerror("Erreur", "Impossible de localiser la plage de données.")
        except KeyError: messagebox.showerror("Erreur", f"La feuille '{SHEET_NAME}' est introuvable.")
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Erreur", f"Une erreur est survenue :\n{type(e).__name__}: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MatriceApp(root)
    root.mainloop()